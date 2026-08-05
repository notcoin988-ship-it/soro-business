"""Генерирует таблицу для менеджера по продажам — заготовку golden set.

Менеджер не должен править YAML: он заполняет три колонки в Excel, а
`scripts/golden_from_xlsx.py` превращает это в `tests/golden.yaml`
(формат раздела 3.1 задач недели 3).

Второй лист — список тем, реально загруженных в базу. Без него вопросы
уедут в темы, которых у бота нет, и метрика получится бессмысленной.

Запуск (из корня soro-business):
    docker compose exec backend python scripts/make_golden_template.py
Файл появится в scripts/golden-set-шаблон.xlsx — оттуда его и отправлять.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, "/code")

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUT = Path("/code/scripts/golden-set-шаблон.xlsx")

EXAMPLES = [
    ("Фоизи амонати «Ояндасоз» чанд аст?", "да", "14,5"),
    ("Какая минимальная сумма вклада «Озод»?", "да", "500"),
    ("Сколько стоит выпуск зарплатной карты?", "да", "бесплатно"),
    ("Какой у меня баланс на карте?", "нет", ""),
    ("Почему у меня списали 90 сомони?", "нет", ""),
    ("А в Амонатбонке ставка по вкладам выше?", "нет", ""),
]

TOPICS = [
    ("Вклады и депозиты", "Ояндасоз · Озод · срочный депозит · Фоиданок · 30 солагии Истиклолият"),
    ("Карты", "Visa Gold · Visa Platinum · Visa Infinite · Visa Virtual · Корти милли"),
    ("Кредиты физлицам", "Манзили · Экспресс · многоцелевой · автокредит · ипотека · онлайн · товары в кредит · Ичтимои"),
    ("Бизнесу", "РКО · зарплатные проекты · торговое финансирование · микрокредиты · агрокредиты · Женщина в бизнесе"),
    ("Сервисы", "Эсхата Мерчант · Эсхата Экспресс · Эсхата Бизнес · Таксии сабз · мобильное приложение"),
    ("Прочее", "денежные переводы · офисы и банкоматы · реквизиты и лицензии · права клиентов"),
    ("Документы", "тарифы физлиц · тарифы юрлиц и ЧП · правила комплексного банковского обслуживания"),
]

HEAD = PatternFill("solid", fgColor="E8506B")  # --rose из прототипа
NOTE = PatternFill("solid", fgColor="FFF4E5")


def main() -> int:
    book = Workbook()

    sheet = book.active
    sheet.title = "Вопросы"

    sheet["A1"] = (
        "Заполните 60 строк: 40 вопросов с ответом в документах (ответ_есть = да) "
        "и 20 без ответа (ответ_есть = нет). Первые шесть строк — образец, их можно "
        "заменить своими."
    )
    sheet["A1"].fill = NOTE
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.merge_cells("A1:C1")
    sheet.row_dimensions[1].height = 45

    headers = ["вопрос", "ответ_есть", "должно_быть_в_ответе"]
    for col, name in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD

    for row, (question, has, must) in enumerate(EXAMPLES, start=3):
        sheet.cell(row=row, column=1, value=question)
        sheet.cell(row=row, column=2, value=has)
        sheet.cell(row=row, column=3, value=must)

    # выпадающий список, чтобы не было «Да», «ДА», «yes» вперемешку
    validation = DataValidation(type="list", formula1='"да,нет"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"B3:B{2 + 60}")

    sheet.column_dimensions["A"].width = 70
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 30
    sheet.freeze_panes = "A3"

    topics = book.create_sheet("Что есть в базе")
    topics["A1"] = (
        "Бот отвечает ТОЛЬКО по этим материалам. Вопрос по теме, которой здесь "
        "нет, автоматически попадает в группу «ответа нет»."
    )
    topics["A1"].fill = NOTE
    topics["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    topics.merge_cells("A1:B1")
    topics.row_dimensions[1].height = 32

    for col, name in enumerate(["раздел", "что загружено"], start=1):
        cell = topics.cell(row=2, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD

    for row, (name, items) in enumerate(TOPICS, start=3):
        topics.cell(row=row, column=1, value=name)
        topics.cell(row=row, column=2, value=items).alignment = Alignment(
            wrap_text=True
        )

    topics.column_dimensions["A"].width = 24
    topics.column_dimensions["B"].width = 90

    book.save(OUT)
    print(f"готово: {OUT}")
    print(f"листов: {len(book.sheetnames)} — {', '.join(book.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
