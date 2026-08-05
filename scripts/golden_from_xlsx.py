"""Таблица от менеджера → `tests/golden.yaml` (формат раздела 3.1).

Менеджер заполняет три колонки в Excel, мы переносим их в YAML. Руками
переносить 60 строк — это гарантированная опечатка в `must_contain`, из-за
которой правильный ответ бота посчитается ошибкой, а порог поедет не туда.

Скрипт заодно проверяет то, что проверить может: пропорцию 40/20,
непустые подсказки у группы «ответ есть», дубли вопросов. Всё, что не
сходится, печатается списком — с этим идти обратно к менеджеру.

Запуск (из корня soro-business):
    docker compose exec backend python scripts/golden_from_xlsx.py
    docker compose exec backend python scripts/golden_from_xlsx.py \\
        --src scripts/golden-set-заполненный.xlsx --out scripts/golden.yaml
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, "/code")

from openpyxl import load_workbook  # noqa: E402

# нормативы раздела 3.1: 40 «ответ есть» + 20 «ответа нет»
NEED_ANSWER = 40
NEED_ESCALATE = 20
YES = {"да", "да ", "yes", "+", "1"}


def read_rows(path: Path) -> list[dict]:
    book = load_workbook(str(path), read_only=True, data_only=True)
    sheet = book["Вопросы"] if "Вопросы" in book.sheetnames else book.worksheets[0]

    rows: list[dict] = []
    for raw in sheet.iter_rows(min_row=3, values_only=True):
        question = (raw[0] or "").strip() if raw and raw[0] else ""
        if not question:
            continue
        has = str(raw[1] or "").strip().lower() if len(raw) > 1 else ""
        must = str(raw[2] or "").strip() if len(raw) > 2 and raw[2] else ""
        rows.append(
            {
                "question": question,
                "expect": "answer" if has in YES else "escalate",
                "must_contain": must,
            }
        )
    book.close()
    return rows


def check(rows: list[dict]) -> list[str]:
    problems: list[str] = []

    answers = [r for r in rows if r["expect"] == "answer"]
    escalates = [r for r in rows if r["expect"] == "escalate"]

    if len(answers) != NEED_ANSWER:
        problems.append(
            f"вопросов с ответом: {len(answers)}, по ТЗ нужно {NEED_ANSWER}"
        )
    if len(escalates) != NEED_ESCALATE:
        problems.append(
            f"вопросов без ответа: {len(escalates)}, по ТЗ нужно {NEED_ESCALATE}"
        )

    for row in answers:
        if not row["must_contain"]:
            problems.append(
                f"нет подсказки «должно_быть_в_ответе»: {row['question'][:60]}"
            )

    seen: set[str] = set()
    for row in rows:
        key = row["question"].lower()
        if key in seen:
            problems.append(f"дубль вопроса: {row['question'][:60]}")
        seen.add(key)

    return problems


def to_yaml(rows: list[dict]) -> str:
    import yaml

    payload = []
    for row in rows:
        item = {"question": row["question"], "expect": row["expect"]}
        if row["must_contain"]:
            item["must_contain"] = row["must_contain"]
        payload.append(item)

    header = (
        "# Golden set — эталонный набор вопросов (раздел 3.1 задач недели 3).\n"
        "#\n"
        "# Собран с менеджером по продажам, преобразован из таблицы скриптом\n"
        "# scripts/golden_from_xlsx.py. Руками не править: правится таблица,\n"
        "# потом перегенерируется файл.\n"
        "#\n"
        "#   expect: answer   — ответ есть в документах, бот обязан его найти\n"
        "#   expect: escalate — ответа нет, бот обязан передать оператору\n"
        "#   must_contain     — подстрока, без которой ответ считается неверным\n"
        "\n"
    )
    return header + yaml.safe_dump(
        payload, allow_unicode=True, sort_keys=False, width=100
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", default="/code/scripts/golden-set-шаблон.xlsx")
    parser.add_argument("--out", default="/code/scripts/golden.yaml")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"нет файла: {src}")
        return 1

    rows = read_rows(src)
    print(f"прочитано строк: {len(rows)}")

    problems = check(rows)
    if problems:
        print(f"\nНЕ СХОДИТСЯ ({len(problems)}):")
        for line in problems:
            print(f"  · {line}")
        print("\nС этим списком — обратно к менеджеру. Файл всё равно записан.")

    Path(args.out).write_text(to_yaml(rows), encoding="utf-8")
    print(f"\nзаписано: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
